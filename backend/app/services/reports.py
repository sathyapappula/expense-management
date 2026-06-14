import csv
import io
from datetime import date
from typing import Optional
from sqlalchemy.orm import Session
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from app.repositories.income import IncomeRepository
from app.repositories.expense import ExpenseRepository


class ReportService:
    def __init__(self, db: Session):
        self.income_repo = IncomeRepository(db)
        self.expense_repo = ExpenseRepository(db)

    def _get_data(
        self,
        user_id: int,
        report_type: str,
        date_from: Optional[date],
        date_to: Optional[date],
    ):
        income_items, _ = self.income_repo.get_by_user(
            user_id, skip=0, limit=10000, date_from=date_from, date_to=date_to
        )
        expense_items, _ = self.expense_repo.get_by_user(
            user_id, skip=0, limit=10000, date_from=date_from, date_to=date_to
        )
        return income_items, expense_items

    def generate_csv(self, user_id: int, report_type: str, date_from: Optional[date], date_to: Optional[date]) -> bytes:
        income_items, expense_items = self._get_data(user_id, report_type, date_from, date_to)
        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow(["=== INCOME ==="])
        writer.writerow(["Date", "Source", "Amount", "Notes"])
        total_income = 0.0
        for item in income_items:
            writer.writerow([item.date, item.source, item.amount, item.notes or ""])
            total_income += item.amount
        writer.writerow(["", "TOTAL", total_income, ""])
        writer.writerow([])

        writer.writerow(["=== EXPENSES ==="])
        writer.writerow(["Date", "Category", "Subcategory", "Amount", "Notes"])
        total_expense = 0.0
        for item in expense_items:
            writer.writerow([item.date, item.category, item.subcategory or "", item.amount, item.notes or ""])
            total_expense += item.amount
        writer.writerow(["", "", "TOTAL", total_expense, ""])
        writer.writerow([])
        writer.writerow(["Net Savings", total_income - total_expense])

        return output.getvalue().encode("utf-8")

    def generate_excel(self, user_id: int, report_type: str, date_from: Optional[date], date_to: Optional[date]) -> bytes:
        income_items, expense_items = self._get_data(user_id, report_type, date_from, date_to)
        wb = Workbook()

        ws_income = wb.active
        ws_income.title = "Income"
        ws_income.append(["Date", "Source", "Amount", "Notes"])
        total_income = 0.0
        for item in income_items:
            ws_income.append([str(item.date), item.source, item.amount, item.notes or ""])
            total_income += item.amount
        ws_income.append(["", "TOTAL", total_income, ""])

        ws_expense = wb.create_sheet("Expenses")
        ws_expense.append(["Date", "Category", "Subcategory", "Amount", "Notes"])
        total_expense = 0.0
        for item in expense_items:
            ws_expense.append([str(item.date), item.category, item.subcategory or "", item.amount, item.notes or ""])
            total_expense += item.amount
        ws_expense.append(["", "", "TOTAL", total_expense, ""])

        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["Metric", "Amount"])
        ws_summary.append(["Total Income", total_income])
        ws_summary.append(["Total Expenses", total_expense])
        ws_summary.append(["Net Savings", total_income - total_expense])

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def generate_pdf(self, user_id: int, report_type: str, date_from: Optional[date], date_to: Optional[date]) -> bytes:
        income_items, expense_items = self._get_data(user_id, report_type, date_from, date_to)
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Personal Finance Report", styles["Title"]))
        elements.append(Spacer(1, 12))

        elements.append(Paragraph("Income", styles["Heading2"]))
        income_data = [["Date", "Source", "Amount", "Notes"]]
        total_income = 0.0
        for item in income_items:
            income_data.append([str(item.date), item.source, f"₹{item.amount:,.2f}", item.notes or ""])
            total_income += item.amount
        income_data.append(["", "TOTAL", f"₹{total_income:,.2f}", ""])
        t = Table(income_data)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.lightgrey),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 12))

        elements.append(Paragraph("Expenses", styles["Heading2"]))
        expense_data = [["Date", "Category", "Amount", "Notes"]]
        total_expense = 0.0
        for item in expense_items:
            expense_data.append([str(item.date), item.category, f"₹{item.amount:,.2f}", item.notes or ""])
            total_expense += item.amount
        expense_data.append(["", "TOTAL", f"₹{total_expense:,.2f}", ""])
        t2 = Table(expense_data)
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.lightgrey),
        ]))
        elements.append(t2)
        elements.append(Spacer(1, 12))

        elements.append(Paragraph("Summary", styles["Heading2"]))
        summary_data = [
            ["Metric", "Amount"],
            ["Total Income", f"₹{total_income:,.2f}"],
            ["Total Expenses", f"₹{total_expense:,.2f}"],
            ["Net Savings", f"₹{total_income - total_expense:,.2f}"],
        ]
        t3 = Table(summary_data)
        t3.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]))
        elements.append(t3)

        doc.build(elements)
        return output.getvalue()
